import openpyxl # for reading excel files
import re, regex, pickle, numpy as np, copy
import heapq, math, random
from sklearn.model_selection import KFold # used for k-fold cross-validation

frequent_terms_num = 20 # removing # of most frequent terms from dictionary
max_results_num = 20 # maximum number of results to show
champions_list_size = 100

ranked_retrieval = True # whether or not to use ranked retrieval
use_index_elimination = True # whether or not to use index elimination technique
use_champions_list = True # whether or not to use champions lists technique
use_heap = True # whether or not to use max-heap in order to return top_k results

arabic_plurals_file = 'arabic_plurals.txt'
verbs_stems_file = 'verbs_stems.txt'

arabic_persian_chars = [['ي', 'ی'], ['ئ', 'ی'], ['ك', 'ک'], ['ة', 'ه'], ['ؤ', 'و'],\
             ['آ', 'ا'], ['إ', 'ا'], ['أ', 'ا'], ['ٱ', 'ا'], ['ء', '']]

# end_words = ['ان', 'ات', 'تر', 'تری', 'ترین', 'م', 'ت', 'ش', 'یی', 'ی', 'ها', 'ا']
end_words = ['ات', 'تر', 'تری', 'ترین', 'یی', 'ی', 'ها']

prefixes = ['ابر', 'باز', 'پاد', 'پارا', 'پسا', 'پیرا', 'ترا', 'فرا', 'هم', 'فرو']
postfixes = ['اسا', 'آگین', 'گین', 'ومند', 'اک', 'اله', 'انه', 'ین'\
    'ینه', 'دان', 'کار' , 'دیس', 'زار', 'سار', 'ستان', 'سرا', 'فام', 'کده', 'گار', \
        'گان', 'گری', 'گر', 'گون', 'لاخ', 'مان', 'مند', 'ناک', 'نده', 'وار', 'واره',\
            'واری', 'ور', 'وش']

past_verb_post = ['م', 'ی', '' , 'یم', 'ید', 'ند']


# Print iterations progress
def print_progress_bar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total: 
        print()

class Document:
    def __init__(self, doc_id, content, topic, url) -> None:
        self.doc_id = doc_id
        self.content = content
        self.topic = topic
        self.url = url

    def __str__(self) -> str:
        return 'doc_id: ' + str(self.doc_id) + '\ttopic: ' + str(self.topic) + '\turl: ' + str(self.url)

    def __repr__(self) -> str:
        return str(self)

class Posting:
    def __init__(self, doc_id, freq) -> None:
        self.doc_id = doc_id
        self.freq = freq
        self.weight = None

    def __str__(self) -> str:
        return 'doc_id: ' + str(self.doc_id) + '\tfreq: ' + str(self.freq) + '\tweight: ' + str(self.weight)

    def __repr__(self) -> str:
        return str(self)


class PostingsList:
    def __init__(self) -> None:
        self.plist = []
        self.term_freq = 0
        self.champs_list = []

    def __str__(self) -> str:
        return 'term_freq: ' + str(self.term_freq) + '\t' + str(self.plist)

    def __repr__(self) -> str:
        return str(self)

    def __lt__(self, other):
        return self.term_freq < other.term_freq

    
    


class IR:
    def __init__(self) -> None:
        self.dictionary = dict()
        self.documents = None
        self.docs_dict = dict()
        self.arabic_plurals_dict = dict()
        self.verbs_dict = dict()
        self.docs_vectors = dict()
        self.clusters = None
        self.cluster_centroids = None
        self.train_docs_vector = dict()

    
    # building the inverted index
    def build_inverted_index(self, file_name):
        global arabic_plurals_file, verbs_stems_file
        self.init_file(file_name)
        # initializing arabic_plurals stemming dictionary
        self.init_arabic_plurals(arabic_plurals_file)
        # initializing verb to verb-stems dictionary
        self.init_verbs_dict(verbs_stems_file)
        indexed_docs_num = 0
        print('Indexing documents...')
        for doc in self.documents:
            print_progress_bar(indexed_docs_num/len(self.documents), 1, prefix = 'Progress:', suffix = 'Complete', length = 50)
            self.index_document(doc)
            indexed_docs_num += 1
        print('Inverted Index Matrix construction completed')
        self.build_docs_dict()
        # removing most frequent items
        self.remove_frequents(frequent_terms_num)
        #calculating tf-idf weights for each posting
        print('Updating Posting Weights...')
        self.update_postings_weights()
        print('Posting Weights Updated!')
        # building champions lists for each term
        print('Building Champions Lists...')
        self.build_champions_lists()
        print('Champions Lists have been built!')
        # building document vectors representations
        print('Building vector-space representations of documents...')
        self.build_doc_vectors()
        print('Vector-space representations of documents have been built!')
        # saving the dictionary
        with open('data\\index.pickle', 'wb') as handle:
            pickle.dump(self.dictionary, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open('data\\docs_dict.pickle', 'wb') as handle:
            pickle.dump(self.docs_dict, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open('data\\arabic_plurals_dict.pickle', 'wb') as handle:
            pickle.dump(self.arabic_plurals_dict, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open('data\\verbs_stems_dict.pickle', 'wb') as handle:
            pickle.dump(self.verbs_dict, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open('data\\docs_vectors.pickle', 'wb') as handle:
            pickle.dump(self.docs_vectors, handle, protocol=pickle.HIGHEST_PROTOCOL)


    # building a dictionary mapping documents IDs to URL
    def build_docs_dict(self):
        for doc in self.documents:
            self.docs_dict[doc.doc_id] = doc.url


    
    # loading the existing inverted index from file
    def load_inverted_index(self):
        with open('data\\index.pickle', 'rb') as handle:
            self.dictionary = pickle.load(handle)
        with open('data\\docs_dict.pickle', 'rb') as handle:
            self.docs_dict = pickle.load(handle)
        with open('data\\arabic_plurals_dict.pickle', 'rb') as handle:
            self.arabic_plurals_dict = pickle.load(handle)
        with open('data\\verbs_stems_dict.pickle', 'rb') as handle:
            self.verbs_dict = pickle.load(handle)
        with open('data\\docs_vectors.pickle', 'rb') as handle:
            self.docs_vectors = pickle.load(handle)

    
    # loading existing clustering data from file
    def load_cluster_data(self):
        with open('data\\clusters.pickle', 'rb') as handle:
            self.clusters = pickle.load(handle)
        with open('data\\cluster_centroids.pickle', 'rb') as handle:
            self.cluster_centroids = pickle.load(handle)



    # initializing documents list by reading the excel dataset
    def init_file(self, file_name):
        print('Initializing documents using Excel file...')
        wb_obj = openpyxl.load_workbook(file_name)
        sheet = wb_obj.active
        headers = []
        dataset = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                for header in row:
                    headers.append(header)
            else:
                document = Document(row[0], row[1], row[2], row[3])
                dataset.append(document)
        self.documents = dataset
        print('Initialized Excel file')

    
    # initializing documents list by reading the excel dataset
    def fetch_documents(self, file_name, trainset=True):
        print('Fetching documents using Excel file...')
        wb_obj = openpyxl.load_workbook(file_name)
        sheet = wb_obj.active
        headers = []
        dataset = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                for header in row:
                    headers.append(header)
            else:
                if trainset:
                    document = Document(row[0], row[1], row[2], row[3])
                else:
                    document = Document(row[0], row[1], None, row[2])
                dataset.append(document)
        print('Fetched all documents')
        return dataset


    # processing the documents one by one for building the index
    def index_document(self, doc):
        tokens = self.get_tokens(doc.content)
        counts = self.get_counts_dict(tokens)
        unique_tokens = counts.keys()
        for unique_token in unique_tokens:
            posting = Posting(doc.doc_id, counts[unique_token])
            self.add_posting(posting, unique_token)

    
    # get tokens for each document
    def get_tokens(self, text):
        persian_letters = r'[آابپتثجچحخدذرزژسشصضطظعغفقکگلمنوهیئؤأإةيكء]+'
        '''
        tokens = re.split('!|,|[|]|\{|\}|\s|-|_|\(|\)|\.|؟|:|»|«|\(|\)|؛|،|\*|&|\
            \^|%|\$|#|@|~|\\|\"|"|\'|;|>|<|\||=|\+|\?', text)
        tokens = list(filter(None, tokens))
        '''
        # tokens = regex.findall(r'[\p{Cf}\p{L}]+', text)
        tokens = re.findall(persian_letters,text)
        # modifying tokens
        # tokens2 = [self.modify_token(token) for token in tokens if self.modify_token(token) != '']
        i = 0
        for token in tokens[:]:
            modified_token = self.modify_token(token)
            if len(modified_token) > 1 and modified_token != '':
                tokens[i] = modified_token
            else:
                del(tokens[i])
                i -= 1
            i += 1
        return tokens

    
    # getting a dictionary of unique terms and the frequencies of each term in a list
    def get_counts_dict(self, tokens):
        counts = dict()
        for token in tokens:
            if token not in counts:
                counts[token] = 1
            else:
                counts[token] += 1
        return counts


    # add posting to the postings_list of the corresponding term in dictionary
    def add_posting(self, posting, term):
        postings_list = None
        if term not in self.dictionary:
            postings_list = PostingsList()
        else:
            postings_list = self.dictionary[term]
        postings_list.term_freq += 1
        postings_list.plist.append(posting)
        self.dictionary[term] = postings_list
        
            

    
    # modify token with stemming, tokenization, normalization, etc
    def modify_token(self, token):
        # to be implemented...
        #
        #
        #
        # m = re.findall(r"^ب[ا-ی]*ید$", token)
        # if m:
        #     if m[0] not in unique_list:
        #         unique_list.append(m[0])
        # removing stop words
        #if len(token) < 3:
        #    return ''
        token = self.normalize(token)
        token = self.stem(token)
        return token


    # normalize the given token
    def normalize(self, token):
        # normalizing characters
        result = token
        # list of [Arabic_character, Persian_character]
        global arabic_persian_chars
        for i in range(len(arabic_persian_chars)):
            if arabic_persian_chars[i][0] in token:
                result = result.replace(arabic_persian_chars[i][0], arabic_persian_chars[i][1])
        '''
        for char_set in arabic_persian_chars:
            result = re.sub(char_set[0], char_set[1], result)
        '''
        return result

    # stemming words and verbs
    def stem(self, token):
        # stemming verbs
        if token in self.verbs_dict:
            token = self.verbs_dict[token]
            return token
        # removing postfixes
        for end in end_words:
            if token.endswith(end):
                token = token[:-len(end)]
        for post in postfixes:
            if token.endswith(post):
                token = token[:-len(post)]
        for pre in prefixes:
            if token.startswith(pre):
                token = token[len(pre):]
        # stemming arabic plurals
        if token in self.arabic_plurals_dict:
            token = self.arabic_plurals_dict[token]
        return token


    # processing queries
    def process_query(self, query):
        tokens = self.get_tokens(query)
        if ranked_retrieval:
            self.process_ranked_query(tokens)
            return
        if len(tokens) == 1:
            self.process_query_single_word(tokens[0])
        elif len(tokens) > 1:
            self.process_query_mult_words(tokens)
        else:
            print('Query is not valid; please make sure your query contains words.')

    
    # processing queries using clusters
    def process_query_using_clustering(self, query, b2_param):
        clusters = self.clusters
        centroids = self.cluster_centroids
        query_vec = self.get_doc_vec(query)
        sim_to_centroids = []
        # comparing query to cluster centroids
        for i in range(len(centroids)):
            sim_to_centroids.append((i, self.cosine_score(query_vec, centroids[i])))
        best_centroids = heapq.nlargest(b2_param, sim_to_centroids, key=lambda p: p[1])
        best_centroids_inds = [v[0] for v in best_centroids]
        cosine_scores = dict()
        # comparing query to the documents of most similar clusters
        for i in range(b2_param):
            cluster = clusters[best_centroids_inds[i]]
            for doc_id, doc_vec in cluster.items():
                score = self.cosine_score(query_vec, doc_vec)
                cosine_scores[doc_id] = score
        top_k = self.retrieve_top_k(cosine_scores, max_results_num)
        self.show_results(top_k)
        return top_k
        
        

    

    # processing queries with only one word
    def process_query_single_word(self, query):
        result_ids = self.get_posting_ids(query)
        if len(result_ids) == 0:
            print("No result found!")
        else:
            showed_result_count = 0
            print('Found ' + str(len(result_ids)) + ' Results: (Showing at max ' + str(max_results_num) + ')')
            for index in result_ids:
                if showed_result_count >= max_results_num:
                    break
                print(index, self.docs_dict[index])
                showed_result_count += 1
        return result_ids


    # processing queries with multiple words - alternative solution using intersection
    def process_query_mult_words_alt(self, tokens):
        id_lists = []
        pointers = []
        result_set = []
        for t in tokens:
            posting_ids = self.get_posting_ids(t)
            id_lists.append(posting_ids)
            if len(posting_ids) > 0:
                pointers.append(0)
            else:
                pointers.append(None)
        terminated = all(x is None for x in pointers)
        while not terminated:
            min_index = len(self.documents) + 1
            min_pointer_ind = len(pointers)
            for i in range(len(pointers)):
                if pointers[i] is None:
                    continue
                id_list = id_lists[i]
                pointer = pointers[i]
                if id_list[pointer] < min_index:
                    min_index = id_list[pointer]
                    min_pointer_ind = i
            # found minimum
            if result_set == [] or result_set[len(result_set)-1][0] != min_index:
                result_set.append([min_index, 1])
            else:
                result_set[len(result_set)-1][1] += 1
            # move min_pointer one step further
            pointers[min_pointer_ind] += 1
            if pointers[min_pointer_ind] >= len(id_lists[min_pointer_ind]):
                pointers[min_pointer_ind] = None
            terminated = all(x is None for x in pointers)
        result_set = sorted(result_set, key=lambda item: item[1], reverse=True)
        return result_set



    # processing queries with multiple words
    def process_query_mult_words(self, tokens):
        id_lists = []
        result_set = dict()
        for t in tokens:
            posting_ids = self.get_posting_ids(t)
            id_lists.append(posting_ids)
        for id_list in id_lists:
            for item in id_list:
                if item not in result_set:
                    result_set[item] = 1
                else:
                    result_set[item] += 1
        result_set = sorted(result_set.items(), key=lambda item: (-item[1], item[0]))
        last_count = 0
        if len(result_set) == 0:
            print("No result found!")
        else:
            print('Found ' + str(len(result_set)) + ' Results: (Showing at max ' + str(max_results_num) + ')')
            showed_results_num = 0
            for item in result_set:
                if showed_results_num >= max_results_num:
                    break
                if last_count != item[1]:
                    print('\nNumber of Words:', item[1])
                    last_count = item[1]
                index = item[0]
                print(index, self.docs_dict[index])
                showed_results_num += 1
        return result_set

    
    # Processing queries using rank-based method
    def process_ranked_query(self, query_tokens):
        query_terms_temp = list(set(query_tokens))
        query_terms = []
        for term in query_terms_temp:
            if term in self.dictionary:
                query_terms.append(term)
        query_terms_freqs = [query_tokens.count(term) for term in query_terms]
        query_terms_weights = []
        # calculating query-terms weights
        for i in range(len(query_terms)):
            term = query_terms[i]
            term_freq = query_terms_freqs[i]
            doc_freq = self.dictionary[term].term_freq
            weight = self.calculate_tf_idf(term_freq, doc_freq, len(self.docs_dict))
            query_terms_weights.append(weight)
        # normalizing query weights
        query_len = sum([weight**2 for weight in query_terms_weights])
        query_len = math.sqrt(query_len)
        query_terms_norm_weights = [weight/query_len for weight in query_terms_weights]
        # calculating each doc's scores
        # scores = [0] * (len(self.docs_dict) + 1)
        scores = dict()
        # doc_lens = [0] * (len(self.docs_dict) + 1)
        docs_lens = dict()
        for term in query_terms:
            posting_ids = None
            if use_champions_list:
                posting_ids = self.get_champs_ids(term)
            else:
                posting_ids = self.get_posting_ids(term)
            for posting_id in posting_ids:
                scores[posting_id] = 0
                docs_lens[posting_id] = 0
        for i in range(len(query_terms)):
            query_term = query_terms[i]
            w_tq = query_terms_norm_weights[i]
            plist = None
            if use_champions_list:
                plist = self.dictionary[query_term].champs_list
            else:
                plist = self.dictionary[query_term].plist
            for posting in plist:
                doc_id, w_td = posting.doc_id, posting.weight
                scores[doc_id] += w_td * w_tq
                docs_lens[doc_id] += w_td**2
        for key in scores.keys():
            if docs_lens[key] != 0:
                docs_lens[key] = math.sqrt(docs_lens[key])
                # normalizing doc-weights vectors by their len in score
                scores[key] /= docs_lens[key]
        top_k = self.retrieve_top_k(scores, max_results_num)
        self.show_results(top_k)
        return top_k

            
    # getting top_k highest cosine scores
    def retrieve_top_k(self, scores, k):
        top_k = []
        if use_heap:
            heap = []
            # building the heap (actually a min heap with -score equivalent max heap with score)
            for item in scores.items():
                heapq.heappush(heap, (-item[1], item[0]))
            # getting k highest scores (equivalent to k smallest -scores)
            for i in range(k):
                if len(heap) == 0:
                    break
                item = heapq.heappop(heap)
                top_k.append([item[1], -item[0]])
        else:
            scores = list(scores.items())
            scores = sorted(scores, key=lambda item: item[1], reverse=True)
            if len(scores) <= k:
                top_k = scores
            else:
                top_k = scores[:k]
        return top_k

    
    # printing query results for the user
    def show_results(self, results):
        result_ids = [res[0] for res in results]
        if len(results) == 0:
            print("No result found!")
        else:
            print('Results: (Showing top ' + str(max_results_num) + ')')
            for index in result_ids:
                print(index, self.docs_dict[index])




    # getting document IDs for a given term
    def get_posting_ids(self, term):
        ids = []
        if term in self.dictionary:
            postings_list = self.dictionary[term]
            postings = postings_list.plist
            for p in postings:
                ids.append(p.doc_id)
        return ids

    
    # getting champions-lists posting ids
    def get_champs_ids(self, term):
        ids = []
        if term in self.dictionary:
            postings_list = self.dictionary[term]
            postings = postings_list.champs_list
            for p in postings:
                ids.append(p.doc_id)
        return ids


    # updating tf-idf weights for every posting
    def update_postings_weights(self):
        for key in self.dictionary.keys():
            plist, doc_freq = self.dictionary[key].plist, self.dictionary[key].term_freq
            for posting in plist:
                term_freq = posting.freq
                posting.weight = self.calculate_tf_idf(term_freq, doc_freq, len(self.docs_dict))


    # building champions lists weights
    def build_champions_lists(self):
        for term in self.dictionary.keys():
            plist = self.dictionary[term].plist
            champs_list = self.dictionary[term].champs_list
            top_postings = heapq.nlargest(champions_list_size, plist, key=lambda p: p.weight)
            top_postings = sorted(top_postings, key=lambda p: p.doc_id)
            for posting in top_postings:
                champs_list.append(posting)

    
    # updating tf-idf weights for a single posting
    def calculate_tf_idf(self, term_freq, doc_freq, num_docs):
        tf_weight = 1 + math.log10(term_freq)
        idf_weight = math.log10(num_docs/doc_freq)
        tf_idf = tf_weight * idf_weight
        return tf_idf
            


    # removing k most frequent term from dictionary
    def remove_frequents(self, k):
        most_frequent = heapq.nlargest(k, self.dictionary, key=self.dictionary.get)
        for key in most_frequent:
            del self.dictionary[key]

    
    # initializing arabic_plurals_dict
    def init_arabic_plurals(self, file_name):
        with open(file_name, 'r', encoding='utf-8') as reader:
            for line in reader:
                words = line.split()
                (singular, plural) = (self.normalize(words[0]), self.normalize(words[1]))
                self.arabic_plurals_dict[plural] = singular


    # generating all verb tenses
    def generate_verb_tenses(self, verb_root):
        # past verbs
        tenses = set()
        past_verb_posts = ['م', 'ی', '' , 'یم', 'ید', 'ند']
        past_root, present_root = verb_root.split(r'$')
        (present_root, past_root) = (self.normalize(present_root), self.normalize(past_root))
        if present_root == 'هست':
            for post in past_verb_posts:
                tenses.add(present_root + post)
                tenses.add('نیست' + post)
            return tenses
        for post in past_verb_posts:
            tenses.add(past_root + post)
            tenses.add('ن' + past_root + post)
            tenses.add('می' + past_root + post)
            tenses.add('نمی' + past_root + post)
        # present verbs
        tenses.add(present_root)
        present_verb_posts = ['م', 'ی', 'د', 'یم', 'ید', 'ند']
        if present_root == 'ا' or present_root == 'گو' or present_root.endswith('ا'):
            present_root = present_root + 'ی'
        for post in present_verb_posts:
            tenses.add(present_root + post)
            tenses.add('می' + present_root + post)
            tenses.add('نمی' + present_root + post)
            # imperatives
            tenses.add('ب' + present_root + post)
            tenses.add('ن' + present_root + post)
        # Masdar
        tenses.add(past_root + 'ن')
        # perfect present
        tenses.add(past_root + 'ه')
        return tenses


    # building dictionary for stemming verbs
    def init_verbs_dict(self, file_name):
        with open(file_name, 'r', encoding='utf-8') as reader:
            for line in reader:
                verb_root = line.split()[0]
                tenses = self.generate_verb_tenses(verb_root)
                for tense in tenses:
                    self.verbs_dict[tense] = verb_root         


    # represent document content as a vector
    def get_doc_vec(self, content):
        tokens = self.get_tokens(content)
        temp_terms = list(set(tokens))
        terms = []
        for term in temp_terms:
            if term in self.dictionary:
                terms.append(term)
        terms_freqs = [tokens.count(term) for term in terms]
        terms_weights = []
        # calculating query-terms weights
        for i in range(len(terms)):
            term = terms[i]
            term_freq = terms_freqs[i]
            doc_freq = self.dictionary[term].term_freq
            weight = self.calculate_tf_idf(term_freq, doc_freq, len(self.docs_dict))
            terms_weights.append(weight)
        # normalizing query weights
        doc_len = sum([weight**2 for weight in terms_weights])
        doc_len = math.sqrt(doc_len)
        terms_norm_weights = [weight/doc_len for weight in terms_weights]
        doc_vec = dict(zip(terms, terms_norm_weights))
        return doc_vec

    # Building vector-space representations of documents
    def build_doc_vectors(self):
        indexed_num = 0
        for doc in self.documents:
            vec = self.get_doc_vec(doc.content)
            self.docs_vectors[doc.doc_id] = vec
            indexed_num += 1
            print_progress_bar(indexed_num/len(self.documents), 1, prefix = 'Progress:', suffix = 'Complete', length = 50)

    
    # Building vector-space representations of train documents
    def build_labeled_doc_vectors(self, dataset):
        indexed_num = 0
        docs_vectors = dict()
        for doc in dataset:
            vec = self.get_doc_vec(doc.content)
            docs_vectors[doc.doc_id] = [vec, doc.topic]
            indexed_num += 1
            print_progress_bar(indexed_num/len(dataset), 1, prefix = 'Progress:', suffix = 'Complete', length = 50)
        return docs_vectors
        

    
    # calculating cosine-similarity between two document vectors
    def cosine_score(self, vec1, vec2):
        score = 0
        base_vec, addit_vec = vec1, vec2
        if len(vec1) > len(vec2):
            base_vec, addit_vec = vec2, vec1
        for term in base_vec.keys():
            if term in addit_vec:
                score += base_vec[term] * addit_vec[term]
        return score

    
    # fiding best (= nearest) centroid for a document
    def find_best_centroid(self, vec, centroids):
        best_ind = 0
        best_score = 0
        for i in range(len(centroids)):
            score = self.cosine_score(vec, centroids[i])
            if score > best_score:
                best_score = score
                best_ind = i
        return best_ind


    # calculate centroid for a given cluster
    def calculate_centroid(self, cluster):
        centroid = dict()
        for vec in cluster.values():
            for term, weight in vec.items():
                if term in centroid:
                    centroid[term] += weight
                else:
                    centroid[term] = weight
        centroid = {k: v / len(cluster) for k, v in centroid.items()}
        # normalizing centroid
        centroid_len = sum([w**2 for w in centroid.values()])
        centroid_len = math.sqrt(centroid_len)
        centroid = {k: v / centroid_len for k, v in centroid.items()}
        return centroid

    
    # checking if clustering has been converged 
    def clustering_converged(self, old_centroids, new_centroids, conv_lim):
        min_sim = 1
        for i in range(len(old_centroids)):
            sim = self.cosine_score(old_centroids[i], new_centroids[i])
            if sim < min_sim:
                min_sim = sim
        if min_sim < conv_lim:
            print('similarity between old and new centroids:', min_sim)
            return False
        print('similarity between old and new centroids:', min_sim)
        print('Converged!')
        return True

    # calculating rss in terms of cosine similarity (the higher -> the better!)
    def calculate_cosine_rss(self, clusters, centroids):
        rss = 0
        for i in range(len(clusters)):
            for vec in clusters[i].values():
                rss += self.cosine_score(vec, centroids[i])
        return rss


    # building clusters using k-means algorithm
    def run_kmeans(self, k):
        conv_lim = 0.98
        max_iter = 100
        seeds_inds = random.sample(range(1, len(self.docs_vectors) + 1), k)
        centroids = [self.docs_vectors[ind] for ind in seeds_inds]
        old_centroids = None
        clusters = [dict() for i in range(k)]
        labels = dict()
        for iter in range(max_iter):
            for i in range(1, len(self.docs_vectors) + 1):
                vec = self.docs_vectors[i]
                label = self.find_best_centroid(vec, centroids)
                labels[i] = label
                clusters[label][i] = vec
            # updating centroids
            old_centroids = copy.deepcopy(centroids)
            for i in range(len(centroids)):
                centroids[i] = self.calculate_centroid(clusters[i])
            # checking convergence
            if self.clustering_converged(old_centroids, centroids, conv_lim):
                break
            print('Iteration:', iter)
        return clusters, centroids, labels


    # running k-means with different initial seeds so as to choose the best clustering
    def cluster(self, num_clusters, num_inits):
        best_clusters = None
        best_centroids = None
        best_labels = None
        best_cosine_rss = None
        for i in range(num_inits):
            clusters, centroids, labels = self.run_kmeans(num_clusters)
            print('Ran k-means for {} out of {} times'.format(i+1, num_inits))
            print('Calculating Cosine-RSS')
            cosine_rss = self.calculate_cosine_rss(clusters, centroids)
            print('Cosine-RSS:', cosine_rss)
            if best_cosine_rss == None or best_cosine_rss < cosine_rss:
                best_cosine_rss = cosine_rss
                best_clusters, best_centroids, best_labels = clusters, centroids, labels
        self.clusters = best_clusters
        self.cluster_centroids = best_centroids
        with open('data\\clusters.pickle', 'wb') as handle:
            pickle.dump(best_clusters, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open('data\\cluster_centroids.pickle', 'wb') as handle:
            pickle.dump(best_centroids, handle, protocol=pickle.HIGHEST_PROTOCOL)

    # fiding k nearest neighbors and returning the prominent label
    def knn_top_topic(self, scores, k):
        k_nearest = heapq.nlargest(k, scores.items(), key=lambda p: p[1][0])
        topics = [v[1][1] for v in k_nearest]
        unique_topics = list(set(topics))
        topic_freqs = [topics.count(t) for t in unique_topics]
        max_freq = max(topic_freqs)
        most_freq_topic = unique_topics[topic_freqs.index(max_freq)]
        return most_freq_topic





    # classifying unlabeled documents using KNN
    def run_knn(self, train_dataset, test_doc_vec, k):
        scores = dict()
        # comparing test vector to every train vector
        for item in train_dataset.items():
            doc_id, doc_vec, doc_label = item[0], item[1][0], item[1][1]
            score = self.cosine_score(test_doc_vec, doc_vec)
            scores[doc_id] = [score, doc_label]
        test_topic = self.knn_top_topic(scores, k)
        return test_topic

    
    # calculate the accuracy of the predicted labels
    def calculate_accuracy(self, real_labels, predicted_labels):
        corrects = 0
        for i in range(len(real_labels)):
            if real_labels[i] == predicted_labels[i]:
                corrects += 1
        return corrects/len(real_labels)



    # evaluate KNN performance on the given test data
    def evaluate_knn(self, train_data, test_data, k_param):
        predicted_labels = []
        real_labels = []
        for test_doc in test_data.items():
            test_id, test_vec, test_label = test_doc[0], test_doc[1][0], test_doc[1][1]
            predicted = self.run_knn(train_data, test_vec, k_param)
            predicted_labels.append(predicted)
            real_labels.append(test_label)
        accuracy = self.calculate_accuracy(real_labels, predicted_labels)
        return accuracy
    
    # finding best K for KNN using k-fold cross-validation
    def cv_best_k(self, dataset, min_k, max_k):
        sample_size = 1000
        keys = random.sample(list(dataset), sample_size)
        values = [dataset[k] for k in keys]
        dataset = dict(zip(keys, values))
        num_folds = 10
        kf10 = KFold(num_folds, shuffle=True)
        dataset_keys = list(dataset.keys())
        mean_accuracies = dict()
        # initializing all accuracies to zero
        for k_param in range(min_k, max_k + 1, 2):
            mean_accuracies[k_param] = 0
        processed = 0
        for train_indices, test_indices in kf10.split(dataset_keys):
            train_doc_ids = [dataset_keys[k] for k in train_indices]
            test_doc_ids = [dataset_keys[k] for k in test_indices]
            train_data = {k: dataset[k] for k in train_doc_ids}
            test_data = {k: dataset[k] for k in test_doc_ids}
            for k_param in range(min_k, max_k + 1, 2):
                calculated_accuracy = self.evaluate_knn(train_data, test_data, k_param)
                mean_accuracies[k_param] += calculated_accuracy
            processed += 1
            print_progress_bar(processed/num_folds, 1, prefix = 'Progress:', suffix = 'Complete', length = 50)
        # normalizing all accuracies to zero
        for k_param in range(min_k, max_k + 1, 2):
            mean_accuracies[k_param] /= num_folds
        best_tup = max(mean_accuracies.items(), key=lambda v: v[1])
        best_k, best_accuracy = best_tup[0], best_tup[1]
        print('Found best k: {}\tAccuracy: {}'.format(best_k, best_accuracy))
        return best_k


    # classifying unlabeled documents by running KNN several times and choosing best k
    def classify(self):
        min_k = 3
        max_k = 19
        train_docs_vects, test_docs_vects = self.load_classification_vectors()
        print('Running 10-fold cross validation...')
        best_k = self.cv_best_k(train_docs_vects, min_k, max_k)
        print('Classifying unlabeled documents...')
        processed = 0
        total_len = len(test_docs_vects.items())
        for test_tup in test_docs_vects.items():
            test_doc_id, test_doc_vec = test_tup[0], test_tup[1][0]
            knn_topic = self.run_knn(train_docs_vects, test_doc_vec, best_k)
            test_tup[1][1] = knn_topic
            print_progress_bar(processed/total_len, 1, prefix = 'Progress:', suffix = 'Complete', length = 50)
            processed += 1
        self.store_classified_data(test_docs_vects)
        

    # build train and test document vectors for classification
    def build_classification_vectors(self, train_dataset_file, test_dataset_file):
        train_dataset = self.fetch_documents(train_dataset_file, trainset=True)
        test_dataset = self.fetch_documents(test_dataset_file, trainset=False)
        train_docs_vects = self.build_labeled_doc_vectors(train_dataset)
        test_docs_vects = self.build_labeled_doc_vectors(test_dataset)
        with open('data\\classification_train_data.pickle', 'wb') as handle:
            pickle.dump(train_docs_vects, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open('data\\classification_test_data.pickle', 'wb') as handle:
            pickle.dump(test_docs_vects, handle, protocol=pickle.HIGHEST_PROTOCOL)

    
    # load classification datasets' vectors
    def load_classification_vectors(self):
        with open('data\\classification_train_data.pickle', 'rb') as handle:
            train_docs_vects = pickle.load(handle)
        with open('data\\classification_test_data.pickle', 'rb') as handle:
            test_docs_vects = pickle.load(handle)
        return train_docs_vects, test_docs_vects
    

    # store classified dataset
    def store_classified_data(self, classified_dataset):
        with open('data\\classified_dataset.pickle', 'wb') as handle:
            pickle.dump(classified_dataset, handle, protocol=pickle.HIGHEST_PROTOCOL)

    
    # load classified dataset
    def load_classified_dataset(self):
        with open('data\\classified_dataset.pickle', 'rb') as handle:
            dataset = pickle.load(handle)
        classified_dataset = dict()
        topics = []
        for doc in dataset.items():
            doc_id, doc_vec, doc_topic = doc[0], doc[1][0], doc[1][1]
            if doc_topic not in topics:
                topics.append(doc_topic)
                topic_dataset = dict()
                topic_dataset[doc_id] = doc_vec
                classified_dataset[doc_topic] = topic_dataset
            else:
                classified_dataset[doc_topic][doc_id] = doc_vec
        return classified_dataset


    # processing queries using clusters
    def process_query_using_classification(self, query):
        topic = re.findall(r'cat:[a-zA-Z]+', query)
        if len(topic) > 0:
            topic = topic[0].replace('cat:', '')
        else:
            topic = None
        classified_dataset = self.load_classified_dataset()
        if topic in classified_dataset.keys():
            topic_docs = classified_dataset[topic]
        else:
            print('Topic does not exist!')
            return
        query_vec = self.get_doc_vec(query)
        cosine_scores = dict()
        # comparing query to the documents of the same topic
        for doc_id, doc_vec in topic_docs.items():
            score = self.cosine_score(query_vec, doc_vec)
            cosine_scores[doc_id] = score
        top_k = self.retrieve_top_k(cosine_scores, max_results_num)
        self.show_results(top_k)
        return top_k
        

    # initializing documents list by reading a list of excel dataset
    def merge_documents(self, file_list, dest_filename):
        print('Fetching documents using Excel file...')
        doc_count = 0
        dataset = []
        for file_name in file_list:
            wb_obj = openpyxl.load_workbook(file_name)
            sheet = wb_obj.active
            headers = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    for header in row:
                        headers.append(header)
                else:
                    doc_count += 1
                    document = Document(doc_count, row[1], row[2], row[3])
                    dataset.append(document)
        print('Fetched all documents')
        print('Storing merged dataset...')
        wb = openpyxl.workbook.Workbook()
        ws1 = wb.active
        ws1.title = "dataset"
        ws1.append(headers)
        for doc in dataset:
            row = [doc.doc_id, doc.content, doc.topic, doc.url]
            ws1.append(row)
        wb.save(filename = dest_filename)
        return dataset